import pandas

# For writing to excel file
import openpyxl

# For writing to google spreadsheet
import gspread
from oauth2client.service_account import ServiceAccountCredentials

COL_OF_TZ = 2
COL_OF_HIT = 8
COL_OF_TREATS = 9
COL_OF_FILENAME = 10

GOOGLE_SHEET_NAME = "Copy of Leumit For Ida"
SHEET_PAGE_NAME = "New Hithayvuyot"


def find_tz_name(cell_value):
    """
    return the t_z, and name in Hebrew
    :param cell_value: the string value
    """
    parts = cell_value.split(" - ")
    name, tz = parts[0], parts[1]
    tz = int(tz.replace("/", ""))
    return tz, name


def register_monthly_ids(file_name):
    """
    Will register this month t_z in dictionary and on computer
    :param file_name: excel file name
    :return:
    """
    # To open the workbook, workbook object is created
    wb_obj = openpyxl.load_workbook(file_name)
    t_z_dict = {}
    for i in ("RONIT", "AVIVA"):
        sheet_obj = wb_obj.get_sheet_by_name(i)
        row, col = 2, COL_OF_TZ
        cell_value = sheet_obj.cell(row, col).value
        while cell_value:
            row += 1
            t_z, name = find_tz_name(cell_value)
            t_z_dict[t_z] = name
            cell_value = sheet_obj.cell(row, col).value

    return t_z_dict


def write_df_to_gspread(data):
    # use creds to create a client to interact with the Google Drive API
    scope = ['https://spreadsheets.google.com/feeds',
             'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('client_secret.json', scope)
    client = gspread.authorize(creds)

    # Find a workbook by name and open the first sheet
    # Make sure you use the right name here.
    sheet = client.open(GOOGLE_SHEET_NAME).worksheet(SHEET_PAGE_NAME)

    list_of_data = [data.columns.values.tolist()] + data.values.tolist()

    for i in range(len(list_of_data)):
        for j in range(len(list_of_data[0])):
            sheet.update_cell(i+1, j+1, list_of_data[i][j])

    print("Done")


def write_data_to_excel(file_name, hit_info):
    wb_obj = openpyxl.load_workbook(file_name)

    for i in ("RONIT", "AVIVA"):
        sheet_obj = wb_obj.get_sheet_by_name(i)
        row, col = sheet_obj.max_row, 1
        cell_value = sheet_obj.cell(row, col).value
        while cell_value > 1:
            row -= 1
            cell_value = sheet_obj.cell(row, col).value

        col = COL_OF_TZ
        cell_value = sheet_obj.cell(row, col).value
        while cell_value:
            t_z, name = find_tz_name(cell_value)
            if t_z in hit_info['t_z']:
                data_row = hit_info[(hit_info['t_z'] == t_z)]
                hits = data_row['hithaybut'].tolist()
                treats = data_row['num_treats'].tolist()
                f_names = data_row['file_name'].tolist()
                # if len(hits) >= 1:
                hit = hits[0]
                num_treats = treats[0]
                f_name = f_names[0]

                sheet_obj.cell(row, COL_OF_HIT).value = hit
                sheet_obj.cell(row, COL_OF_TREATS).value = num_treats
                sheet_obj.cell(row, COL_OF_FILENAME).value = f_name

            row += 1
            cell_value = sheet_obj.cell(row, col).value

    wb_obj.save(file_name)


hit_info = pandas.DataFrame.from_csv("Hithayvuyot_06012020.xlsx")
write_data_to_excel("MayHit.xlsx", hit_info)
write_df_to_gspread(hit_info)
