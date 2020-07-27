# for gmail API
from __future__ import print_function
import os.path
# for downloading the pdfs
from download_atachments import downloading_recipet_pdfs
# for extracting text from pdf
import pdf2image
import os
from PIL import Image
import pytesseract
# for saving the data
import pandas
# for requesting the input
import re
# for writing to google spreadsheet
import gspread
from oauth2client.service_account import ServiceAccountCredentials

import openpyxl

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

T_Z = "ת.ז"
CODE_TIPUL = "קוד טיפול"
TIPUL = "טיפול"
HITHAYVUT = "התחייבות"

EMPTY = ""

GOOGLE_SHEET_NAME = "Copy of Leumit For Ida"
SHEET_PAGE_NAME = "New Hithayvuyot"

WARNING_COLOR = '\033[93m'
END_COLOR = '\033[0m'

BAD_FORMAT_MSG = "Bad format. Enter again:\n"

COL_OF_TZ = 2
COL_OF_HIT = 8
COL_OF_TREATS = 9
COL_OF_FILENAME = 10


def main_function():
    """
    This function manages the flow of the program. it requests the user dates to
    check, dowloads the data, extracts the data to data frame and at the end
    writes it to a google sheet excel
    """
    date_to_read_from, date_to_read_to = request_dates_from_user()
    query = "073-7619000"

    downloading_recipet_pdfs(query, date_to_read_from, date_to_read_to)

    hit_info = pandas.DataFrame(extracting_info_from_files())
    hit_info.to_excel("Hithayvuyot_" + date_to_read_from.replace("/", "") + ".xlsx",
                      engine='xlsxwriter')

    # hit_info = pandas.DataFrame.from_csv("Hithayvuyot_06012020.xlsx")
    # write_data_to_excel("MayHit.xlsx", hit_info)

    # write_df_to_gspread(hit_info)


def write_data_to_excel(file_name, hit_info):
    wb_obj = openpyxl.load_workbook(file_name)

    for i in ("RONIT", "AVIVA"):
        sheet_obj = wb_obj.get_sheet_by_name(i)
        row, col = sheet_obj.max_row, 1
        cell_value = sheet_obj.cell(row, col).value
        while cell_value > 1:
            row -= 1
            cell_value = sheet_obj.cell(row, col).value

        col = COL_OF_TZ
        cell_value = sheet_obj.cell(row, col).value
        while cell_value:
            t_z, name = find_tz_name(cell_value)
            if t_z in hit_info['t_z']:
                data_row = hit_info[(hit_info['t_z'] == t_z)]
                hits = data_row['hithaybut'].tolist()
                treats = data_row['num_treats'].tolist()
                f_names = data_row['file_name'].tolist()
                # if len(hits) >= 1:
                hit = hits[0]
                num_treats = treats[0]
                f_name = f_names[0]

                sheet_obj.cell(row, COL_OF_HIT).value = hit
                sheet_obj.cell(row, COL_OF_TREATS).value = num_treats
                sheet_obj.cell(row, COL_OF_FILENAME).value = f_name

            row += 1
            cell_value = sheet_obj.cell(row, col).value

    wb_obj.save(file_name)


def find_tz_name(cell_value):
    """
    return the t_z, and name in Hebrew
    :param cell_value: the string value
    """
    parts = cell_value.split(" - ")
    name, tz = parts[0], parts[1]
    tz = int(tz.replace("/", ""))
    return tz, name


def register_monthly_ids(file_name):
    """
    Will register this month t_z in dictionary and on computer
    :param file_name: excel file name
    :return:
    """
    # To open the workbook, workbook object is created
    wb_obj = openpyxl.load_workbook(file_name)
    t_z_dict = {}
    for i in ("RONIT", "AVIVA"):
        sheet_obj = wb_obj.get_sheet_by_name(i)
        row, col = 2, COL_OF_TZ
        cell_value = sheet_obj.cell(row, col).value
        while cell_value:
            row += 1
            t_z, name = find_tz_name(cell_value)
            t_z_dict[t_z] = name
            cell_value = sheet_obj.cell(row, col).value

    return t_z_dict


def write_df_to_gspread(data):
    # use creds to create a client to interact with the Google Drive API
    scope = ['https://spreadsheets.google.com/feeds',
             'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('client_secret.json', scope)
    client = gspread.authorize(creds)

    # Find a workbook by name and open the first sheet
    # Make sure you use the right name here.
    sheet = client.open(GOOGLE_SHEET_NAME).worksheet(SHEET_PAGE_NAME)

    list_of_data = [data.columns.values.tolist()] + data.values.tolist()

    for i in range(len(list_of_data)):
        for j in range(len(list_of_data[0])):
            sheet.update_cell(i+1, j+1, list_of_data[i][j])

    print("Done")


def handle_line_with_tz(line):
    """
    Given a line text from the pdf, trying to find the id.
    We know of 2 possible lines in the pdf where id can be found, in the first option
    we can possibly find the number of hithayvut
    :param line:
    :return: a 4 element tuple - t_z, if found_tz, hithayvut, if found_hit
    if not found, there are empty values in t_z/hithayvut accordingly
    """
    parts = line.split(" ")
    found_tz, found_hit = False, False
    hithayvut = EMPTY
    if line.startswith(T_Z) or parts[1] == T_Z:
        t_z = parts[3]
        hithayvut = parts[-1]

        if t_z.isdigit():
            found_tz = True
        if hithayvut.isdigit():
            found_hit = True

    else:
        # the t_z is the last one
        t_z = parts[-1]
        if t_z.isdigit():
            found_tz = True
        else:
            t_z = EMPTY

    return t_z, found_tz, hithayvut, found_hit


def handle_line_with_hit(line):
    """
    Given a line text from the pdf, trying to find the hithayvut
    :param line:
    :return: a 2 element tuple - hithayvut, if found_hit
    if not found, there is a dummy value in hithayvut
    """
    parts = line.split(" ")
    found_hit = False
    hithayvut = parts[-1]
    if hithayvut.isdigit():
        found_hit = True

    return hithayvut, found_hit


def handle_line_with_num_tipulim(line):
    """
    Given a line text from the pdf, trying to find the number of treatments
    :param line:
    :return: num_treatments, if it's not a number it will be an empty string value
    """
    parts = line.split(" ")
    num_treats = parts[-1]
    if not num_treats.isdigit():
        num_treats = EMPTY

    return num_treats


def parse_text_and_get_info(text_from_file):
    """
    Parsing the text from file in a way we'll fins the t.z., hithayvut, and # tipulim
    :param text_from_file:
    :return: t.z., hithayvut, and # tipulim
    """
    t_z, hithayvut, num_treats = EMPTY, EMPTY, EMPTY
    lines = text_from_file.split("\n")
    found_tz = False
    found_hit = False
    tipul_in_next_line = False

    for line in lines:
        if tipul_in_next_line and TIPUL in line:
            num_treats = handle_line_with_num_tipulim(line)
            break

        elif not found_tz and T_Z in line:
            t_z, found_tz, hit, found_hit_here = handle_line_with_tz(line)
            if not found_hit and found_hit_here:
                hithayvut = hit
                found_hit = True

        elif not found_hit and line.startswith(HITHAYVUT):
            hithayvut, found_hit = handle_line_with_hit(line)

        elif CODE_TIPUL in line:
            tipul_in_next_line = True

    return t_z, hithayvut, num_treats


def extract_text(file_path, image_name):
    """
    given a file, extracting the text from it
    :return:
    """
    images = pdf2image.convert_from_path(file_path)
    for img in images:
        img.save(image_name + ".jpeg", 'JPEG')

    pytesseract.pytesseract.tesseract_cmd = \
        r'C:\Program Files\Tesseract-OCR\tesseract.exe'

    return pytesseract.image_to_string(Image.open(image_name + ".jpeg"), lang='heb')


def print_whats_missing(new_filename, t_z, hithayvut, num_treats):
    """
    Checks which on of the details is missing and prints as an Error
    """
    print(WARNING_COLOR + "!!! Error in " + new_filename + ": ")
    if t_z == EMPTY:
        print("t_z is missing")
    if hithayvut == EMPTY:
        print("# hithayvut is missing")
    if num_treats == EMPTY:
        print("# of treatments is missing")
    print(END_COLOR)


def check_validity(t_z, hithayvut, num_treats):
    """
    Goes over the t_z that exists in the system this month and make sure it's available
    :param t_z: id
    :param hithayvut: num of the hit.
    :param num_treats: number of treatments
    :return: the updated values for the data
    """
    if t_z.isdigit():
        t_z = int(t_z)
        if not t_z in tz_dict:
            t_z = "Error"

    if hithayvut.isdigit():
        hithayvut = int(hithayvut)
    else:
        hithayvut = "Error"

    if num_treats.isdigit():
        num_treats = int(num_treats)
    else:
        num_treats = "Error"

    return t_z, hithayvut, num_treats


def extracting_info_from_files():
    """
    Goes over the pdf's in this directory and extracts the relevant information: t.z,
    number of recipet, and number of treatments. writes it down in to an excel sheet
    """
    counter = 119
    data = pandas.DataFrame(columns=['t_z', 'hithayvut', 'num_treats', 'file_name'])
    for filename in os.listdir("Attachments"):
        if filename.endswith(".pdf"):
            image_path = "Attachments\\file" + str(counter)
            file_path = os.path.join("Attachments", filename)

            text_from_file = extract_text(file_path, image_path)
            t_z, hithayvut, num_treats = parse_text_and_get_info(text_from_file)

            new_filename = image_path + ".pdf"
            os.rename(file_path, new_filename)

            t_z, hithayvut, num_treats = check_validity(t_z, hithayvut, num_treats)
            data.loc[counter] = [t_z, hithayvut, num_treats, image_path + ".pdf"]

            if t_z == EMPTY or hithayvut == EMPTY or num_treats == EMPTY:
                print_whats_missing(new_filename, t_z, hithayvut, num_treats)

            counter += 1
    return data


def check_input_match(regex_to_match, msg_to_present):
    """
    :param regex_to_match: the regex case needed
    :param msg_to_present:
    :return: date part that matches the regex and contains digits only
    """
    date_part = input(msg_to_present)
    while not re.match(regex_to_match, date_part) or not date_part.isdigit():
        date_part = input(BAD_FORMAT_MSG)
    return date_part


def request_dates_from_user():
    """
    Requests the user a start date and end date
    :return: start date, end date
    """
    print("Hello Yos! \nPlease enter the START date for your mails - ")

    day = check_input_match(r"[0-9][0-9]", "Enter day 01-31\n")
    month = check_input_match(r"[0-9][0-9]", "Enter month 01-12\n")
    year = check_input_match(r"[2][0-9]", "Enter year 20-25\n")
    full_year = "20" + year
    start_date = month + "/" + day + "/" + full_year
    print("Your start day is: ", start_date)

    print("Please enter the END date for your mails -\nif it's today just press ENTER")
    day = input("Enter day or ENTER\n")
    if day == "":
        return start_date, EMPTY

    day = check_input_match(r"[0-9][0-9]", "Enter day 01-31\n")
    month = check_input_match(r"[0-9][0-9]", "Enter month 01-12\n")
    end_date = month + "/" + day + "/" + full_year
    print("Your start day is: ", end_date)

    return start_date, end_date


if __name__ == '__main__':
    """
    Main program runs first the ID check, then the main function
    """
    
    # Enter a file name with the t.z.
    tz_dict = register_monthly_ids("MayHit.xlsx")
    main_function()
